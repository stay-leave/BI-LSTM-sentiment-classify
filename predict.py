#coding='utf-8'
import numpy as np
from dataset import Data_convert#引入数据集
from model import BI_lstm#模型
from config import ModelConfig#配置
import openpyxl
import xlwt
import torch


def predict(config, model, pred_loader):
    #调用训练好的模型对新句子进行预测,以分好词，编码的形式（调用dataset
    #model.eval()
    pred_all=[]#所有结果
    with torch.no_grad():
        #模型初始化赋值
        h = model.init_hidden(config.batch_size_pred)#根据待预测的句子数确定
        for dat,id in pred_loader:
            h = tuple([each.data for each in h])
            #dat=torch.Tensor(dat)#列表转张量
            dat=dat.cuda()#GPU
            #print('dat的数据：')
            #print(dat)
            output,h= model(dat, h)#输出
            #print('output的数据：')
            #print(output)
            #pred=output.detach().numpy()#转换数据时不需要保留梯度信息
            pred=output.cpu().numpy().tolist()#输出列表[0.521,0.465,...]
            pred_all=pred_all+pred
            #最后匹配的时候需要与输入的TXT文件列表做同时循环
    return pred_all

def save_file(config, alls):
    """保存结果到excel
    """
    f = openpyxl.Workbook()
    sheet1 = f.create_sheet('sheet1')
    sheet1['A1'] = 'id'
    sheet1['B1'] = '评论内容'
    sheet1['C1'] = '情感值'
    sheet1['D1'] = '情感类别'# [0,0.5]负向，(0.5,1]正向

    i = 2  # openpyxl最小值是1，写入的是xlsx
    for all in alls:  # 遍历每一页
        # for data in all:#遍历每一行
        for j in range(1, len(all) + 1):  # 取每一单元格
            # sheet1.write(i,j,all[j])#写入单元格
            sheet1.cell(row=i, column=j, value=all[j - 1])
        i = i + 1  # 往下一行
    f.save(config.save_pred_path)

if __name__ == '__main__':
    #进行预测
    config=ModelConfig()#实例化配置
    #读取字典
    vocab_train = np.load(config.save_dict_path+'.npy',allow_pickle=True)
    vocab_train= vocab_train.item()
    #GPU
    device = torch.device('cuda') if torch.cuda.is_available() else torch.device("cpu")
    #加载模型
    new_model = torch.load(config.save_model_path)#加载训练好的模型
    new_model.to(device)
    #预测数据加载。实例化类，下面要用到它的方法
    data_pred=Data_convert(config.input_path_pred,config.seq_len,config.batch_size_pred)#根据待预测的句子数确定
    datas=data_pred.txt_file(config.input_path_pred)#读入
    datas,id=data_pred.splitt_nos(datas)#取句子（未分词），列表,j是id
    data_token=[]#分词数组
    for i in datas:
        data_token.append(data_pred.tokenlize(i))
    
    '''
    data=[]#数字序列列表
    for i in datas:
        #i=data_pred.splitt(i)#分词
        i=data_pred.seq_to_array(i,vocab_train)#转为数字序列
        data.append(i)#此时可以进行predict
    #data=torch.Tensor(data)#列表转张量
    '''

    pred_loader=data_pred.data_for_pred_txt(data_token,vocab_train,id)#加载到tensor
    #运行模型，输出结果
    pred=predict(config,new_model,pred_loader)
    #保存结果
    all=[]#数组
    for i,d,k in zip(id,datas,pred):#将原始句子和预测值同时循环
        #打印输出
        k=round(k,5)#保留五位小数
        #print('句子'+d+'的情感值为：'+str(k))
        if k>=0.5:
            #print('句子'+d+'为：正向情感')
            s='正向'
        else:
            #print('句子'+d+'为：负向情感')
            s='负向'
        all.append([i,d,k,s])
    save_file(config, all)#保存结果
    #需要根据预测的句子数量确定
